from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from .models import Payroll
import openpyxl

@login_required
def payroll_list(request):
    employee = request.user.employee
    
    if employee.is_admin:
        payrolls = Payroll.objects.all()
    else:
        payrolls = Payroll.objects.filter(employee=employee)
    
    return render(request, 'payroll/payroll_list.html', {'payrolls': payrolls})

@login_required
def payslip_view(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    
    # 本人または管理者のみ閲覧可能
    if not (request.user.employee.is_admin or payroll.employee == request.user.employee):
        return HttpResponse('閲覧権限がありません', status=403)
    
    return render(request, 'payroll/payslip.html', {'payroll': payroll})

@login_required
def payslip_download(request, pk):
    payroll = get_object_or_404(Payroll, pk=pk)
    
    if not (request.user.employee.is_admin or payroll.employee == request.user.employee):
        return HttpResponse('ダウンロード権限がありません', status=403)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '給与明細'
    
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from decimal import Decimal
    
    title_font = Font(name='メイリオ', size=16, bold=True)
    header_font = Font(name='メイリオ', size=11, bold=True)
    normal_font = Font(name='メイリオ', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # タイトル
    ws.merge_cells('A1:F1')
    ws['A1'] = '給与明細書'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 社員情報
    ws['A3'] = '社員番号'
    ws['B3'] = payroll.employee.employee_no
    ws['A4'] = '氏名'
    ws['B4'] = payroll.employee.name
    ws['A5'] = '部署'
    ws['B5'] = payroll.employee.department
    
    ws['E3'] = '対象月'
    ws['F3'] = payroll.month
    ws['E4'] = '発行日'
    ws['F4'] = payroll.created_at.strftime('%Y-%m-%d')
    
    # 支給項目
    ws['A7'] = '支給項目'
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws.merge_cells('A7:B7')
    
    ws['A8'] = '項目'
    ws['B8'] = '金額'
    for cell in ['A8', 'B8']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].border = thin_border
    
    row = 9
    items = [
        ('基本給', payroll.base_salary),
        ('残業代', payroll.overtime_pay),
    ]
    if payroll.bonus > 0:
        items.append(('賞与', payroll.bonus))
    
    for item, amount in items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = float(amount)
        ws[f'B{row}'].number_format = '¥#,##0'
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    ws[f'A{row}'] = '支給合計'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = total_fill
    ws[f'B{row}'] = float(payroll.base_salary + payroll.overtime_pay + payroll.bonus)
    ws[f'B{row}'].number_format = '¥#,##0'
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].fill = total_fill
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border
    
    # 控除項目
    row += 2
    ws[f'D7'] = '控除項目'
    ws[f'D7'].font = header_font
    ws[f'D7'].fill = header_fill
    ws.merge_cells('D7:E7')
    
    ws['D8'] = '項目'
    ws['E8'] = '金額'
    for cell in ['D8', 'E8']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].border = thin_border
    
    deduction_row = 9
    deductions = [
        ('健康保険料', payroll.health_insurance),
        ('厚生年金保険料', payroll.pension_insurance),
        ('雇用保険料', payroll.employment_insurance),
        ('所得税', payroll.income_tax),
        ('住民税', payroll.residence_tax),
    ]
    if payroll.other_deduction > 0:
        deductions.append(('その他控除', payroll.other_deduction))
    
    for item, amount in deductions:
        ws[f'D{deduction_row}'] = item
        ws[f'E{deduction_row}'] = float(amount)
        ws[f'E{deduction_row}'].number_format = '¥#,##0'
        ws[f'D{deduction_row}'].border = thin_border
        ws[f'E{deduction_row}'].border = thin_border
        deduction_row += 1
    
    ws[f'D{deduction_row}'] = '控除合計'
    ws[f'D{deduction_row}'].font = header_font
    ws[f'D{deduction_row}'].fill = total_fill
    ws[f'E{deduction_row}'] = float(payroll.total_deduction)
    ws[f'E{deduction_row}'].number_format = '¥#,##0'
    ws[f'E{deduction_row}'].font = header_font
    ws[f'E{deduction_row}'].fill = total_fill
    ws[f'D{deduction_row}'].border = thin_border
    ws[f'E{deduction_row}'].border = thin_border
    
    # 差引支給額
    row = max(row, deduction_row) + 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = '差引支給額'
    ws[f'A{row}'].font = Font(name='メイリオ', size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].border = thin_border
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = float(payroll.total_salary)
    ws[f'A{row}'].number_format = '¥#,##0'
    ws[f'A{row}'].font = Font(name='メイリオ', size=18, bold=True, color='C00000')
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 35
    
    # 列幅調整
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payslip_{payroll.month}_{payroll.employee.employee_no}.xlsx'
    wb.save(response)
    return response